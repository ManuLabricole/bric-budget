"""
demo/generators.py — rend un profil (#118) au format EXACT de chaque banque.

Déterministe (Random seedé) : même `anchor` → mêmes fichiers. Les fichiers produits
passent par les VRAIS connecteurs (round-trip testé dans tests/demo/). 0 donnée réelle.

Fichiers supportés :
    ubs_checking / ubs_savings → CSV UBS (;, utf-8-sig, IBAN ligne 2)
    yuh                        → CSV Yuh (;, utf-8-sig, noms entre quotes)
    cic                        → XLSX CIC (multi-feuilles, RIB par feuille, compte EUR)
"""

from __future__ import annotations

import calendar
import csv
import datetime
import io
from datetime import date
from pathlib import Path
from random import Random

import openpyxl

from demo import profiles
from demo.profiles import Flow

# Graine fixe → données reproductibles (utile pour les fixtures committées stables).
_SEED = 4242

# En-têtes exacts attendus par les connecteurs (cf. fixtures de test).
_UBS_META_BLANK_THEN_HEADER = [
    "Date de transaction",
    "Heure de transaction",
    "Date de comptabilisation",
    "Date de valeur",
    "Monnaie",
    "Débit",
    "Crédit",
    "Sous-montant",
    "Solde",
    "No de transaction",
    "Description1",
    "Description2",
    "Description3",
    "Notes de bas de page",
    "",  # colonne finale vide → ligne se termine par ';'
]
_YUH_HEADER = [
    "DATE",
    "ACTIVITY TYPE",
    "ACTIVITY NAME",
    "DEBIT",
    "DEBIT CURRENCY",
    "CREDIT",
    "CREDIT CURRENCY",
    "CARD NUMBER",
    "LOCALITY",
    "RECIPIENT",
    "SENDER",
    "FEES/COMMISSION",
    "BUY/SELL",
    "QUANTITY",
    "ASSET",
    "PRICE PER UNIT",
]

# En-têtes de la ligne 5 (HEADER_ROW) d'une feuille compte CIC.
_CIC_HEADER = ["Date", "Valeur", "Libellé", "Débit", "Crédit", "Solde", "Dev"]


# ── Déroulé temporel ──────────────────────────────────────────────────────────


def _months_back(anchor: date, months: int) -> list[tuple[int, int]]:
    """(year, month) des `months` derniers mois jusqu'à anchor inclus, ordre croissant."""
    out: list[tuple[int, int]] = []
    y, m = anchor.year, anchor.month
    for _ in range(months):
        out.append((y, m))
        m -= 1
        if m == 0:
            y, m = y - 1, 12
    return list(reversed(out))


def _day_in_month(rng: Random, preferred: int, year: int, month: int) -> date:
    """Jour dans le mois : preferred ±3 (clampé 1..28), ou aléatoire si preferred=0."""
    last = calendar.monthrange(year, month)[1]
    if preferred == 0:
        day = rng.randint(1, last)
    else:
        day = max(1, min(preferred + rng.randint(-3, 3), min(28, last)))
    return date(year, month, day)


def _jitter(rng: Random, amount: float, variance: float = 0.12) -> float:
    """Variance ±variance% pour simuler le monde réel."""
    return round(amount * (1 + rng.uniform(-variance, variance)), 2)


def _events(
    flows: list[Flow], *, months: int, anchor: date, rng: Random
) -> list[tuple[date, Flow, float]]:
    """Déroule les flux sur la période → (date, flow, montant jitté), triés par date."""
    events: list[tuple[date, Flow, float]] = []
    for year, month in _months_back(anchor, months):
        for flow in flows:
            # Occasionnel : présent ~60% des mois seulement.
            if not flow.recurrent and rng.random() > 0.6:
                continue
            d = _day_in_month(rng, flow.day, year, month)
            events.append((d, flow, _jitter(rng, flow.amount)))
    events.sort(key=lambda e: e[0])
    return events


def _csv_string(rows: list[list[str]]) -> str:
    """Écrit des lignes en CSV ';' avec terminateur '\\n' (comme les exports réels)."""
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";", lineterminator="\n")
    writer.writerows(rows)
    return buf.getvalue()


# ── Générateurs par banque ────────────────────────────────────────────────────


def generate_ubs_csv(
    flows: list[Flow],
    *,
    iban: str,
    account_number: str,
    ref_prefix: str,
    months: int,
    anchor: date,
    initial_balance: float = 0.0,
    rng: Random | None = None,
) -> str:
    """Rend un relevé UBS : 8 lignes méta (IBAN ligne 2) + ligne vide + en-tête + transactions.

    ref_prefix : préfixe des « No de transaction ». DOIT être unique par compte —
    l'import_hash UBS en dérive, donc deux comptes qui réutilisent TX0001 verraient
    leurs transactions dédupliquées entre eux (l'import_hash est global).
    """
    rng = rng or Random(_SEED)
    events = _events(flows, months=months, anchor=anchor, rng=rng)

    balance = initial_balance
    data_rows: list[list[str]] = []
    for i, (d, flow, amount) in enumerate(events, start=1):
        signed = amount if flow.direction == "credit" else -amount
        balance = round(balance + signed, 2)
        debit = f"{signed:.2f}" if signed < 0 else ""
        credit = f"{signed:.2f}" if signed > 0 else ""
        desc2 = "Virement entrant" if flow.direction == "credit" else "Ordre e-banking"
        data_rows.append(
            [
                d.isoformat(),
                "",  # heure
                d.isoformat(),
                d.isoformat(),
                "CHF",
                debit,
                credit,
                "",  # sous-montant
                f"{balance:.2f}",
                f"{ref_prefix}{i:04d}",
                flow.label,  # Description1 = marchand
                desc2,
                f"Ref {ref_prefix}{i:04d}",
                "",  # notes
                "",  # colonne finale → trailing ';'
            ]
        )

    date_min = events[0][0] if events else anchor
    date_max = events[-1][0] if events else anchor
    meta = [
        ["Numéro de compte:", account_number, ""],
        ["IBAN:", iban, ""],
        ["Du:", date_min.isoformat(), ""],
        ["Au:", date_max.isoformat(), ""],
        ["Solde initial:", f"{initial_balance:.2f}", ""],
        ["Solde final:", f"{balance:.2f}", ""],
        ["Évaluation en:", "CHF", ""],
        ["Nombre de transactions dans cette période:", str(len(data_rows)), ""],
        [],  # ligne vide avant l'en-tête (ligne 9)
        _UBS_META_BLANK_THEN_HEADER,
    ]
    return _csv_string(meta + data_rows)


def generate_yuh_csv(
    flows: list[Flow],
    *,
    card_last_four: str,
    months: int,
    anchor: date,
    rng: Random | None = None,
) -> str:
    """Rend un relevé Yuh : en-tête + transactions carte (noms entre quotes comme l'export réel)."""
    rng = rng or Random(_SEED)
    events = _events(flows, months=months, anchor=anchor, rng=rng)

    data_rows: list[list[str]] = []
    for d, flow, amount in events:
        # Yuh entoure les libellés de guillemets ; csv.writer les échappe en "...".
        quoted = f'"{flow.label}"'
        debit = f"{-amount:.2f}" if flow.direction == "debit" else ""
        credit = f"{amount:.2f}" if flow.direction == "credit" else ""
        data_rows.append(
            [
                d.strftime("%d/%m/%Y"),
                "CARD_TRANSACTION_OUT",
                quoted,
                debit,
                "CHF",
                credit,
                "CHF" if credit else "",
                f"**** {card_last_four}" if card_last_four else "",
                "Genève",
                quoted,  # RECIPIENT
                "",  # SENDER
                "0",  # FEES/COMMISSION
                "",  # BUY/SELL
                "",  # QUANTITY
                "",  # ASSET
                "",  # PRICE PER UNIT
            ]
        )
    return _csv_string([_YUH_HEADER] + data_rows)


def _cic_title(account_type: str, date_max: date) -> str:
    """Titre de la ligne A1 cohérent avec account_type — doit contenir (EUR) et le
    marqueur que `_detect_account_type` cherche (C/C → checking, sinon savings)."""
    suffix = f"(EUR) au {date_max:%d/%m/%Y}"
    if account_type == "checking":
        # "C/C" déclenche checking dans le parser (_detect_account_type).
        return f"Situation de votre compte C/C CONTRAT PERSONNEL GLOBAL {suffix}"
    # Tout titre sans "C/C" ni "CONTRAT PERSONNEL" → savings.
    return f"Situation de votre compte LIVRET A SUP {suffix}"


def generate_cic_xlsx(
    sheets: list[tuple[str, str, str, list[Flow]]],
    *,
    anchor: date,
    months: int,
    rng: Random | None = None,
) -> bytes:
    """Rend un classeur CIC : feuille 'Vos comptes' (résumé minimal) + 1 feuille par
    compte. EUR. Retourne les bytes du .xlsx (openpyxl + BytesIO).

    sheets : liste de (sheet_name, rib_spaced, account_type, flows). Le RIB doit être
    DISTINCT par feuille — il entre dans l'import_hash CIC, donc deux feuilles au même
    RIB verraient leurs transactions dédupliquées l'une contre l'autre.

    Chaque feuille reproduit le layout du vrai export CIC (cf. CICConnector) :
        A1 = titre (contient (EUR) + marqueur de type), A2 = "R.I.B. : <rib espacé>",
        A3/A4 = libellés, ligne 5 = en-têtes, lignes 6+ = transactions
        (Date=datetime.datetime, Débit négatif|None, Crédit positif|None, Solde courant,
        Dev="EUR"), puis une ligne footer (col D = "Solde au … : ", col F = solde final).
    """
    rng = rng or Random(_SEED)
    wb = openpyxl.Workbook()

    # Feuille récapitulative — ignorée par le parser, mais "Vos comptes" est requise
    # par matches_file() pour reconnaître le fichier comme un export CIC.
    ws_summary = wb.active
    ws_summary.title = "Vos comptes"
    ws_summary["A1"] = "Résumé de vos comptes"

    for sheet_name, rib_spaced, account_type, flows in sheets:
        events = _events(flows, months=months, anchor=anchor, rng=rng)

        # Solde courant : on déroule depuis 0 puis on rebase pour que le solde final
        # tombe sur une valeur synthétique stable (les chiffres réels n'importent pas).
        balance = 0.0
        rows: list[list] = []
        for d, flow, amount in events:
            signed = amount if flow.direction == "credit" else -amount
            balance = round(balance + signed, 2)
            debit = -amount if flow.direction == "debit" else None
            credit = amount if flow.direction == "credit" else None
            rows.append(
                [
                    # Le parser fait hasattr(date_val, "date") → exige un datetime,
                    # PAS un date (date n'a pas de .date()). D'où datetime.datetime.
                    datetime.datetime(d.year, d.month, d.day),
                    datetime.datetime(d.year, d.month, d.day),  # Valeur (ignorée)
                    flow.label,
                    debit,
                    credit,
                    balance,  # Solde courant après l'opération (colonne F)
                    "EUR",
                ]
            )

        date_max = events[-1][0] if events else anchor

        ws = wb.create_sheet(sheet_name)
        ws["A1"] = _cic_title(account_type, date_max)
        ws["A2"] = f"R.I.B. : {rib_spaced}"
        ws["A3"] = "Solde initial"
        ws["A4"] = "Détails des opérations"
        # Le 1er append() atterrit en ligne 5 (max_row=4 après A4) → en-têtes,
        # puis lignes 6+ = transactions, comme l'attend DATA_START_ROW=6.
        ws.append(_CIC_HEADER)
        for row in rows:
            ws.append(row)
        # Footer lu par _extract_balance_from_sheet : col D (index 3) = "Solde au … : ",
        # col F (index 5) = solde final. Col A = None → skippée comme transaction.
        ws.append(
            [None, None, None, f"Solde au {date_max:%d/%m/%Y} : ", None, balance, None]
        )

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ── Écriture sur disque (encodage + nom de fichier corrects) ──────────────────

# bank → (fonction de rendu, nom de fichier). Les CSV sont lus en utf-8-sig ;
# CIC est un .xlsx binaire (bytes), pas de texte/BOM.
_BANKS = ("ubs_checking", "ubs_savings", "yuh", "yuh_savings", "cic")

# Flux de démonstration CIC (compte EUR fictif) — inline car spécifiques au format
# CIC et sans équivalent dans profiles.py (persona UBS/Yuh en CHF).
_CIC_CHECKING_FLOWS: list[Flow] = [
    Flow("VIR SEPA EMPLOYER SA SALAIRE", 2500, 25, "credit"),
    Flow("PRLV SEPA LOYER RESIDENCE", 850, 1, "debit"),
    Flow("PAIEMENT PSC 0306 PARIS MONOPRIX CARTE 8703", 70, 6, "debit"),
    Flow("PAIEMENT CB 1206 LYON SNCF WEB CARTE 8703", 45, 12, "debit"),
    Flow("PRLV SEPA ELECTRICITE EDF", 60, 10, "debit"),
]
_CIC_SAVINGS_FLOWS: list[Flow] = [
    Flow("VIR SEPA EPARGNE MENSUELLE", 300, 26, "credit"),
]
# RIB synthétiques DISTINCTS par feuille (SR-008 : espacés → hook anti-IBAN OK ;
# que des zéros sauf le dernier chiffre → import_hash disjoints). PUBLICS car le
# seeder en a besoin pour créer les comptes CIC (Account.contract_number).
CIC_CHECKING_RIB = "00000 00000 00000000001"
CIC_SAVINGS_RIB = "00000 00000 00000000002"


def write_bank_file(
    bank: str,
    dest_dir: Path,
    *,
    months: int = 12,
    anchor: date | None = None,
    rng: Random | None = None,
) -> Path:
    """Génère le fichier `bank` dans dest_dir (utf-8-sig) et retourne son chemin.

    Utilisé par le seeder (anchor=today → dates fraîches) ET par dev_generate_fixtures
    (anchor fixe → fixtures committées stables).
    """
    anchor = anchor or date.today()
    rng = rng or Random(_SEED)
    dest_dir.mkdir(parents=True, exist_ok=True)

    if bank == "ubs_checking":
        content = generate_ubs_csv(
            profiles.UBS_CHECKING_FLOWS,
            iban=profiles.DEMO_UBS_CHECKING_IBAN,
            account_number=profiles.DEMO_UBS_CHECKING_NUMBER,
            ref_prefix="UBSC",
            months=months,
            anchor=anchor,
            rng=rng,
        )
        path = dest_dir / "ubs_checking_demo.csv"
    elif bank == "ubs_savings":
        content = generate_ubs_csv(
            profiles.UBS_SAVINGS_FLOWS,
            iban=profiles.DEMO_UBS_SAVINGS_IBAN,
            account_number=profiles.DEMO_UBS_SAVINGS_NUMBER,
            ref_prefix="UBSS",
            months=months,
            anchor=anchor,
            rng=rng,
        )
        path = dest_dir / "ubs_savings_demo.csv"
    elif bank == "yuh":
        content = generate_yuh_csv(
            profiles.YUH_CARD_FLOWS,
            card_last_four=profiles.DEMO_YUH_CARD_LAST_FOUR,
            months=months,
            anchor=anchor,
            rng=rng,
        )
        path = dest_dir / "yuh_demo.csv"
    elif bank == "yuh_savings":
        # Épargne Yuh : versements mensuels entrants, sans carte.
        content = generate_yuh_csv(
            profiles.YUH_SAVINGS_FLOWS,
            card_last_four="",
            months=months,
            anchor=anchor,
            rng=rng,
        )
        path = dest_dir / "yuh_savings_demo.csv"
    elif bank == "cic":
        # CIC = .xlsx binaire, multi-feuilles (1 par compte) → write_bytes, pas de BOM.
        data = generate_cic_xlsx(
            [
                ("Cpt courant", CIC_CHECKING_RIB, "checking", _CIC_CHECKING_FLOWS),
                ("Cpt livret", CIC_SAVINGS_RIB, "savings", _CIC_SAVINGS_FLOWS),
            ],
            anchor=anchor,
            months=months,
            rng=rng,
        )
        path = dest_dir / "cic_demo.xlsx"
        path.write_bytes(data)
        return path
    else:
        raise ValueError(f"Banque démo inconnue : {bank!r} (attendu : {_BANKS})")

    # utf-8-sig : les connecteurs UBS/Yuh lisent avec le BOM (comme les vrais exports).
    path.write_text(content, encoding="utf-8-sig")
    return path


# bank → sous-dossier des fixtures committées (demo/fixtures/<subdir>/).
_FIXTURE_SUBDIR = {
    "ubs_checking": "ubs",
    "ubs_savings": "ubs",
    "yuh": "yuh",
    "yuh_savings": "yuh",
    "cic": "cic",
}


def write_fixtures(root: Path, *, anchor: date, months: int = 12) -> list[Path]:
    """(Re)génère toutes les fixtures committées sous root/<banque>/.

    Appelée par dev_generate_fixtures avec un anchor FIXE → fichiers stables
    (pas de churn git à chaque regénération).
    """
    return [
        write_bank_file(bank, root / subdir, months=months, anchor=anchor)
        for bank, subdir in _FIXTURE_SUBDIR.items()
    ]
