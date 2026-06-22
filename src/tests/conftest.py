"""
tests/conftest.py — Fixtures partagées entre tous les sous-dossiers de tests.

Pourquoi ce fichier existe-t-il à ce niveau (src/tests/) et pas dans un sous-dossier ?
----------------------------------------------------------------------------------------
pytest charge les conftest.py hiérarchiquement, du plus proche vers la racine.
Un conftest.py dans src/tests/connectors/ n'est PAS visible depuis src/tests/integration/.

La fixture cic_file est utilisée à la fois dans :
    - src/tests/connectors/test_cic.py  (tests unitaires du parser)
    - src/tests/integration/test_import_integration.py  (test E2E)

Elle doit donc vivre ici, au niveau parent commun des deux sous-dossiers.

Règle de rangement :
    - Fixtures utilisées dans UN seul sous-dossier → conftest.py local du sous-dossier
    - Fixtures utilisées dans PLUSIEURS sous-dossiers → ce fichier
"""

from datetime import datetime
from pathlib import Path

import openpyxl
import pytest

from services import logos as _logos_service


@pytest.fixture(autouse=True)
def _no_network_logo_fetch(monkeypatch):
    """
    Garde réseau global : aucun test ne doit télécharger un logo pour de vrai
    (le post_save Institution déclenche fetch_logo dès qu'un domain est posé).
    fetch_logo attrape l'exception par contrat → comportement « échec réseau »,
    silencieux et sans effet. Les tests du service re-patchent _download eux-mêmes.
    """

    def _blocked(url: str, dest: Path) -> None:
        raise OSError("réseau désactivé dans les tests (garde conftest)")

    monkeypatch.setattr(_logos_service, "_download", _blocked)


@pytest.fixture(autouse=True)
def _reset_icon_map_cache():
    """
    Isolation des tests — vide le cache d'icônes entre chaque test (issue #192).

    `get_institution_icon_map()` mémoïse son résultat dans un lru_cache
    PROCESS-GLOBAL (`services.logos._icon_map_cached`) qui survit à toute la
    session pytest. Sans reset, un test qui peuple/invalide ce cache (logos,
    logo_repair, backfill_logos) contamine l'état lu par les tests suivants :
    source de non-déterminisme (flake d'un test IDOR observé ~1 run sur 14).

    On vide AVANT (pas de fuite entrante d'un test précédent) ET APRÈS (les
    chemins de fichiers tmp_path de ce test ne fuient pas vers les suivants).
    """
    _logos_service.clear_institution_icon_cache()
    yield
    _logos_service.clear_institution_icon_cache()


@pytest.fixture
def cic_file(tmp_path) -> Path:
    """
    Crée un fichier Excel CIC minimal avec 2 feuilles de compte :
    - "Cpt CC"       : compte courant, 3 transactions EUR, RIB 10096 18027 00064764601
    - "Cpt LivretA"  : livret épargne, 2 transactions EUR, RIB 10096 18027 00087654321
    - "Vos comptes"  : feuille récapitulative (ignorée par le parser)

    Structure d'une feuille CIC :
        Row 1 : titre (contient "C/C" ou "LIVRET")
        Row 2 : RIB "R.I.B. : XXXXX XXXXX XXXXXXXXXXX"
        Row 3 : libellé solde initial
        Row 4 : libellé section
        Row 5 : en-têtes colonnes (Date, Valeur, Libellé, Débit, Crédit, Solde, Dev)
        Row 6+ : transactions (datetime, datetime, str, float|None, float|None, float, str)
        Dernier : ligne footer "Solde au DD/MM/YYYY : " en colonne D, valeur en F

    tmp_path est un fixture pytest built-in : crée un dossier temporaire unique par test,
    supprimé automatiquement après. Le .xlsx y est sauvegardé pour que CICConnector puisse
    l'ouvrir depuis le disque (openpyxl ne lit pas depuis un BytesIO en mode data_only).
    """
    wb = openpyxl.Workbook()

    # ── Feuille récapitulative (ignorée par matches_file et parser) ─────────────
    ws_summary = wb.active
    ws_summary.title = "Vos comptes"
    ws_summary["A1"] = "Résumé de vos comptes"

    # ── Feuille C/C (compte courant) ────────────────────────────────────────────
    ws_cc = wb.create_sheet("Cpt CC")
    ws_cc["A1"] = (
        "Situation de votre compte C/C CONTRAT PERSONNEL GLOBAL (EUR) au 31/03/2026"
    )
    ws_cc["A2"] = "R.I.B. : 10096 18027 00064764601"
    ws_cc["A3"] = "Solde initial"
    ws_cc["A4"] = "Détails des opérations"
    # Row 5 : en-têtes (le parser DATA_START_ROW=6 commence après)
    ws_cc.append(["Date", "Valeur", "Libellé", "Débit", "Crédit", "Solde", "Dev"])
    # Row 6 : paiement carte (débit négatif, card last four dans libellé)
    ws_cc.append(
        [
            datetime(2026, 3, 17),
            datetime(2026, 3, 17),
            "PAIEMENT PSC 1703 LAUSANNE MIGROS CARTE 8703",
            -25.40,
            None,
            974.60,
            "EUR",
        ]
    )
    # Row 7 : virement entrant salaire (crédit positif)
    ws_cc.append(
        [
            datetime(2026, 3, 15),
            datetime(2026, 3, 15),
            "VIR SEPA EMPLOYER SA",
            None,
            2500.00,
            1000.00,
            "EUR",
        ]
    )
    # Row 8 : paiement CB (débit, description avec préfixe CB)
    ws_cc.append(
        [
            datetime(2026, 3, 10),
            datetime(2026, 3, 10),
            "PAIEMENT CB 1003 PARIS SNCF WEB MOBILE CARTE 8703",
            -45.00,
            None,
            500.00,
            "EUR",
        ]
    )
    # Row 9 : footer — col D = string "Solde au...", col F = solde final
    # date (col A) = None → skippé par "if date_val is None: continue"
    ws_cc.append([None, None, None, "Solde au 31/03/2026 : ", None, 798.27, None])

    # ── Feuille Livret A (compte épargne) ───────────────────────────────────────
    ws_la = wb.create_sheet("Cpt LivretA")
    ws_la["A1"] = "Situation de votre compte LIVRET A SUP (EUR) au 31/03/2026"
    ws_la["A2"] = "R.I.B. : 10096 18027 00087654321"
    ws_la["A3"] = "Solde initial"
    ws_la["A4"] = "Détails des opérations"
    ws_la.append(["Date", "Valeur", "Libellé", "Débit", "Crédit", "Solde", "Dev"])
    # Row 6 : virement épargne mensuel
    ws_la.append(
        [
            datetime(2026, 3, 1),
            datetime(2026, 3, 1),
            "VIR SEPA EPARGNE MENSUELLE",
            None,
            300.00,
            5300.00,
            "EUR",
        ]
    )
    # Row 7 : intérêts annuels
    ws_la.append(
        [
            datetime(2026, 2, 28),
            datetime(2026, 2, 28),
            "INTERETS ANNUELS",
            None,
            15.00,
            5000.00,
            "EUR",
        ]
    )
    # Row 8 : footer
    ws_la.append([None, None, None, "Solde au 31/03/2026 : ", None, 5315.00, None])

    filepath: Path = tmp_path / "cic_export.xlsx"
    wb.save(str(filepath))
    return filepath
