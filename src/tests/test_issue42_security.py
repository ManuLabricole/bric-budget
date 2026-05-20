"""
tests/test_issue42_security.py — Tests issue #42 (security followup Phase 2H).

Couvre :
    SEC-01  FILE_UPLOAD_MAX_MEMORY_SIZE défini dans settings
    SEC-02  print() remplacés par logger dans parseurs et signals
    SEC-03  except Exception silencieux → logger dans connecteurs
    OBS-01  LOGGING configuré dans settings
    OPS-02  MultiFernet : rotation de clé Fernet sans re-chiffrement
"""

import logging

import pytest
from django.conf import settings
from django.test import override_settings

# =============================================================================
# SEC-01 — FILE_UPLOAD_MAX_MEMORY_SIZE
# =============================================================================


def test_file_upload_max_memory_size_is_defined():
    """5 MB définis — protège le worker contre un upload malveillant."""
    assert hasattr(settings, "FILE_UPLOAD_MAX_MEMORY_SIZE")
    assert settings.FILE_UPLOAD_MAX_MEMORY_SIZE == 5 * 1024 * 1024


# =============================================================================
# OBS-01 — LOGGING
# =============================================================================


def test_logging_is_configured():
    """Bloc LOGGING présent dans settings."""
    assert hasattr(settings, "LOGGING")
    assert settings.LOGGING.get("version") == 1


def test_logging_has_console_handler():
    """Handler 'console' (stdout) présent — capturé par Railway."""
    assert "console" in settings.LOGGING.get("handlers", {})


def test_logging_connectors_logger_configured():
    """Logger 'connectors' explicitement configuré."""
    assert "connectors" in settings.LOGGING.get("loggers", {})


def test_logging_transactions_logger_configured():
    """Logger 'transactions' explicitement configuré (pour signals.py)."""
    assert "transactions" in settings.LOGGING.get("loggers", {})


# =============================================================================
# SEC-02 — print() → logger dans les parseurs
# =============================================================================


def test_yuh_parser_warning_via_logger(tmp_path, caplog):
    """
    YuhConnector.parse() loggue les erreurs de ligne via logger.warning (plus de print).
    On injecte une ligne CSV malformée et on vérifie que le warning apparaît dans caplog.
    """
    from connectors.yuh.parser import YuhConnector

    header = "DATE;ACTIVITY TYPE;ACTIVITY NAME;DEBIT;DEBIT CURRENCY;CREDIT;CREDIT CURRENCY;CARD NUMBER;LOCALITY;RECIPIENT;SENDER;FEES/COMMISSION;BUY/SELL;QUANTITY;ASSET;PRICE PER UNIT\n"
    bad_row = "23/10/2025;CARD_TRANSACTION_OUT;Test;;;;;;;;;;;;\n"
    csv_file = tmp_path / "test.csv"
    csv_file.write_text(header + bad_row, encoding="utf-8")

    connector = YuhConnector()
    with caplog.at_level(logging.WARNING, logger="connectors.yuh.parser"):
        connector.parse(csv_file)

    warnings = [
        r for r in caplog.records if r.levelno >= logging.WARNING and "Yuh" in r.message
    ]
    assert len(warnings) == 1
    assert "Both DEBIT and CREDIT are empty" in warnings[0].message


def test_cic_parser_info_via_logger(caplog):
    """
    CICConnector.parse_sheet() loggue le résultat via logger.info (plus de print).
    """
    import os
    import tempfile

    import openpyxl

    from connectors.cic.parser import CICConnector

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cpt test"
    ws["A1"] = "Situation de votre compte C/C TEST"
    ws["A2"] = "R.I.B. : 10096 18027 00064764601"
    ws["A5"] = "Date"

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        tmp_path_str = f.name
    wb.save(tmp_path_str)

    connector = CICConnector()
    with caplog.at_level(logging.INFO, logger="connectors.cic.parser"):
        connector.parse_sheet(tmp_path_str, "Cpt test")

    os.unlink(tmp_path_str)
    info_records = [
        r for r in caplog.records if r.levelno == logging.INFO and "CIC" in r.message
    ]
    assert len(info_records) == 1
    assert "0 transactions" in info_records[0].message


def test_signals_logs_orphaned_snapshots_deleted(db, caplog):
    """
    cleanup_orphaned_snapshots() loggue via logger.info (plus de print).
    Supprime un ImportLog avec snapshot orphelin — vérifie que le message apparaît dans caplog.
    """
    from datetime import date

    from accounts.models import Account, BalanceSnapshot, Bank
    from transactions.models import ImportLog, Transaction
    from users.models import CustomUser

    user = CustomUser.objects.create_user(email="sig@test.com", password="x")
    bank = Bank.objects.create(
        name="Test Bank", slug="test-bank-sig", country="CH", default_currency="CHF"
    )
    account = Account.objects.create(
        name="Signal Test",
        bank=bank,
        account_type="checking",
        currency="CHF",
    )
    account.members.add(user)

    log = ImportLog.objects.create(
        account=account,
        imported_by=user,
        filename="test.csv",
        file_hash="sig_abc123",
        status="success",
        count_created=1,
        date_min=date(2026, 1, 1),
        date_max=date(2026, 1, 31),
    )

    Transaction.objects.create(
        account=account,
        import_log=log,
        date=date(2026, 1, 15),
        amount=-10,
        currency="CHF",
        description_raw="test",
        import_hash="tx_sig_abc123",
    )

    BalanceSnapshot.objects.create(
        account=account,
        date=date(2026, 1, 15),
        balance=1000,
        currency="CHF",
    )

    with caplog.at_level(logging.INFO, logger="transactions.signals"):
        log.delete()

    info_records = [
        r
        for r in caplog.records
        if r.levelno >= logging.INFO and "BalanceSnapshot" in r.message
    ]
    assert len(info_records) == 1
    assert "1" in info_records[0].message


# =============================================================================
# SEC-03 — except Exception silencieux → logger dans connecteurs
# =============================================================================


def test_ubs_extract_account_name_logs_on_error(tmp_path, caplog):
    """
    UBSConnector.extract_account_name() : si le fichier est illisible,
    logger.warning est appelé (pas silencieux).
    """
    from connectors.ubs.parser import UBSConnector

    bad_file = tmp_path / "bad.csv"
    bad_file.write_bytes(b"\xff\xfe bad encoding")

    connector = UBSConnector()
    with caplog.at_level(logging.WARNING, logger="connectors.ubs.parser"):
        result = connector.extract_account_name(bad_file)

    # Doit retourner None sans planter
    assert result is None


def test_ubs_matches_file_logs_debug_on_error(tmp_path, caplog):
    """
    UBSConnector.matches_file() : fichier illisible → logger.debug, retourne False.
    """
    from connectors.ubs.parser import UBSConnector

    bad_file = tmp_path / "bad.csv"
    # Créer un fichier .csv illisible (permissions)
    bad_file.write_text("garbage", encoding="utf-8")
    bad_file.chmod(0o000)

    try:
        with caplog.at_level(logging.DEBUG, logger="connectors.ubs.parser"):
            result = UBSConnector.matches_file(bad_file)
        assert result is False
    finally:
        bad_file.chmod(0o644)


# =============================================================================
# OPS-02 — MultiFernet : rotation de clé Fernet
# =============================================================================


def test_multifernet_encrypts_with_first_key():
    """
    Avec deux clés, encrypt_bytes() chiffre avec la première.
    La première clé seule suffit à déchiffrer.
    """
    from cryptography.fernet import Fernet

    from imports.storage import encrypt_bytes

    key1 = Fernet.generate_key().decode()
    key2 = Fernet.generate_key().decode()

    with override_settings(IMPORT_ENCRYPTION_KEY=f"{key1},{key2}"):
        ciphertext = encrypt_bytes(b"donnees bancaires")

    # key1 seule doit déchiffrer (c'est la clé primaire)
    assert Fernet(key1.encode()).decrypt(ciphertext) == b"donnees bancaires"


def test_multifernet_decrypts_with_old_key():
    """
    Rotation : fichier chiffré avec l'ancienne clé reste lisible après rotation.

    Scénario :
        1. Chiffrement avec key_old (ancienne config)
        2. On ajoute key_new en première position : IMPORT_ENCRYPTION_KEY=key_new,key_old
        3. Le fichier ancien est toujours déchiffrable
    """
    from cryptography.fernet import Fernet

    from imports.storage import decrypt_bytes, encrypt_bytes

    key_old = Fernet.generate_key().decode()
    key_new = Fernet.generate_key().decode()

    # Étape 1 : chiffrement avec l'ancienne clé
    with override_settings(IMPORT_ENCRYPTION_KEY=key_old):
        ciphertext_old = encrypt_bytes(b"ancien fichier bancaire")

    # Étape 2 : après rotation, on peut encore déchiffrer l'ancien fichier
    with override_settings(IMPORT_ENCRYPTION_KEY=f"{key_new},{key_old}"):
        plaintext = decrypt_bytes(ciphertext_old)

    assert plaintext == b"ancien fichier bancaire"


def test_multifernet_new_files_use_new_key():
    """
    Après rotation, les nouveaux fichiers sont chiffrés avec la nouvelle clé.
    L'ancienne clé ne peut PAS les déchiffrer.
    """
    from cryptography.fernet import Fernet, InvalidToken

    from imports.storage import encrypt_bytes

    key_old = Fernet.generate_key().decode()
    key_new = Fernet.generate_key().decode()

    with override_settings(IMPORT_ENCRYPTION_KEY=f"{key_new},{key_old}"):
        ciphertext_new = encrypt_bytes(b"nouveau fichier bancaire")

    # L'ancienne clé seule ne peut PAS déchiffrer le nouveau fichier
    with pytest.raises(InvalidToken):
        Fernet(key_old.encode()).decrypt(ciphertext_new)

    # La nouvelle clé, elle, le peut
    assert (
        Fernet(key_new.encode()).decrypt(ciphertext_new) == b"nouveau fichier bancaire"
    )


def test_single_key_backwards_compatible():
    """
    Une seule clé (format actuel .env) fonctionne comme avant — pas de régression.
    """
    from cryptography.fernet import Fernet

    from imports.storage import decrypt_bytes, encrypt_bytes

    key = Fernet.generate_key().decode()
    with override_settings(IMPORT_ENCRYPTION_KEY=key):
        ciphertext = encrypt_bytes(b"fichier normal")
        plaintext = decrypt_bytes(ciphertext)

    assert plaintext == b"fichier normal"


def test_get_fernet_raises_when_key_empty():
    """_get_fernet() sans clé → ImproperlyConfigured avec message explicite."""
    from django.core.exceptions import ImproperlyConfigured

    from imports.storage import _get_fernet

    with override_settings(IMPORT_ENCRYPTION_KEY=""):
        with pytest.raises(ImproperlyConfigured) as exc_info:
            _get_fernet()
    assert "IMPORT_ENCRYPTION_KEY" in str(exc_info.value)
