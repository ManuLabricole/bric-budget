"""
connectors/cic/parser.py — CIC France Excel parser.

File format
-----------
Extension : .xlsx
Library   : openpyxl (data_only=True — reads cached values, no formula engine needed)

Sheet structure
---------------
One Excel file = multiple accounts (checking + savings on same export).
Relevant sheets : all sheets whose name starts with "Cpt " — one per account.
Ignored sheets  : "Vos comptes" (summary), "hidden_data", "hidden" (metadata).

Per-sheet layout
----------------
Row 1 : title       → "Situation de votre compte C/C CONTRAT ... (EUR) au 30/03/2026"
Row 2 : RIB         → "R.I.B. : 10096 18027 00064764601"   ← account identifier
Row 3 : initial balance label
Row 4 : section header
Row 5 : column headers → Date · Valeur · Libellé · Débit · Crédit · Solde · Dev
Row 6+: data rows
Last rows : empty rows + footer row "Solde au DD/MM/YYYY : " + section header

Column mapping (0-indexed when iterating, 1-indexed in Excel)
---------+------+------------------------------------------------------------
A (0)    | Date | openpyxl returns datetime.datetime directly — no conversion
B (1)    | Valeur (value date) — ignored
C (2)    | Libellé — raw description
D (3)    | Débit — negative float if present, None otherwise
E (4)    | Crédit — positive float if present, None otherwise
F (5)    | Solde — running balance (often None per row, present in footer)
G (6)    | Dev — currency code (EUR)

Quirks
------
- openpyxl reads dates as datetime.datetime (data_only=True handles this)
- Footer row: column D contains "Solde au DD/MM/YYYY : " — detect by isinstance(str)
- Some rows at the end are fully empty — skip rows where date (col A) is None
- Card number in description: "PAIEMENT PSC 3003 ... CARTE 8703" → last_four="8703"
- No time column (CIC doesn't export transaction time)
- Both debit and credit use float with dot as decimal separator

Account matching
----------------
Each sheet has a RIB in row 2: "R.I.B. : 10096 18027 00064764601"
We normalise (strip spaces, remove "R.I.B. : ") and match against:
  - Account.contract_number → the resolver matches every CIC sheet on the RIB
    (C/C and Livret alike) ; IBAN n'est PAS utilisé pour CIC.

Account type detection (from title row 1):
  "C/C" in title → checking
  "LIVRET" or "DEVELOPPEMENT DURABLE" → savings
"""

import hashlib
import logging
import re
from pathlib import Path

import openpyxl

from connectors.base import BaseConnector, TransactionDict

logger = logging.getLogger(__name__)

# Sheets to skip — not account transaction sheets
NON_ACCOUNT_SHEETS = {"Vos comptes", "hidden_data", "hidden"}

# Row numbers (1-based, as in Excel)
HEADER_ROW = 5  # column names row
DATA_START_ROW = 6  # first transaction row


class CICConnector(BaseConnector):
    """
    Parses CIC France Excel exports into normalized TransactionDicts.

    A single CIC export file contains multiple sheets — one per account.
    Use get_account_sheets() to discover all accounts, then parse_sheet()
    for each one individually. The management command handles the loop.

    The inherited parse() method parses all account sheets combined.
    """

    # =========================================================================
    # Public API
    # =========================================================================

    @classmethod
    def matches_file(cls, filepath: Path) -> bool:
        """
        Return True if this file looks like a CIC Excel export.

        Fast check: must be .xlsx and contain a "Vos comptes" sheet.
        We don't open the whole file — just peek at the sheet names.
        """
        if filepath.suffix.lower() != ".xlsx":
            return False
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            return "Vos comptes" in sheet_names
        except Exception:
            logger.debug("[CIC] matches_file failed for %s", filepath, exc_info=True)
            return False

    def get_account_sheets(self, filepath: Path) -> list[dict]:
        """
        Return info about each account sheet in the file.

        Returns a list of dicts, one per account sheet:
        {
            "sheet_name" : "Cpt 18027 00064764601",
            "rib"        : "10096XXXXXXXXXXXXXXXXXXX",   # normalised (no spaces)
            "rib_raw"    : "10096 18027 00064764601", # as in the file
            "balance"    : 798.27,                    # from footer
            "account_type_hint": "checking",          # or "savings"
        }
        """
        wb = openpyxl.load_workbook(filepath, data_only=True)
        result = []

        for sheet_name in wb.sheetnames:
            if sheet_name in NON_ACCOUNT_SHEETS:
                continue
            ws = wb[sheet_name]
            rib_raw = self._extract_rib(ws)
            rib = rib_raw.replace(" ", "") if rib_raw else None
            balance = self._extract_balance_from_sheet(ws)
            account_type_hint = self._detect_account_type(ws)
            result.append(
                {
                    "sheet_name": sheet_name,
                    "rib": rib,
                    "rib_raw": rib_raw,
                    "balance": balance,
                    "account_type_hint": account_type_hint,
                }
            )

        wb.close()
        return result

    def parse_sheet(self, filepath: Path, sheet_name: str) -> list[TransactionDict]:
        """
        Parse all transactions from a single account sheet.

        Skips header rows (1-4), the column header row (5), empty rows,
        and the footer row ("Solde au...").
        """
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb[sheet_name]

        # Extract RIB once — used in import_hash to ensure uniqueness across accounts
        rib = (self._extract_rib(ws) or sheet_name).replace(" ", "")

        transactions = []
        skipped = 0
        # occurrence_counters : même mécanisme que Yuh.
        # Clé = (rib, date_str, amount, description_raw) → index d'occurrence.
        # Permet de distinguer deux transactions CIC strictement identiques le même jour
        # (ex: deux tickets RATP à 2.10€) sans dépendre de la position dans le fichier.
        occurrence_counters: dict[tuple, int] = {}

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=DATA_START_ROW, values_only=True), start=DATA_START_ROW
        ):
            date_val, _, libelle, debit, credit, solde, currency = (
                row[0],
                row[1],
                row[2],
                row[3],
                row[4],
                row[5],  # Solde — running balance after this transaction
                row[6],
            )

            # Skip empty rows (no date)
            if date_val is None:
                continue

            # Skip footer row — its "date" column contains a string like "Solde au..."
            # openpyxl returns a datetime for real date cells; footer rows have None in col A
            # Sometimes footer detection relies on column D being a string
            if isinstance(debit, str) and "Solde au" in debit:
                continue

            # Skip rows that look like section headers (col A is a string, not a date)
            if not hasattr(date_val, "date"):
                continue

            try:
                transactions.append(
                    self._parse_row(
                        date_val,
                        libelle,
                        debit,
                        credit,
                        solde,
                        currency,
                        rib,
                        row_idx,
                        occurrence_counters,
                    )
                )
            except Exception as e:
                logger.warning("[CIC] row %d: %s — skipped", row_idx, e, exc_info=True)
                skipped += 1

        wb.close()
        logger.info(
            "[CIC] Sheet '%s': %d transactions, %d skipped",
            sheet_name,
            len(transactions),
            skipped,
        )
        return transactions

    def parse(self, filepath: Path) -> list[TransactionDict]:
        """
        Parse all account sheets in the file and return combined transactions.

        Implements the BaseConnector contract. For per-account reporting,
        use get_account_sheets() + parse_sheet() directly in the command.
        """
        all_txs = []
        for sheet_info in self.get_account_sheets(filepath):
            all_txs.extend(self.parse_sheet(filepath, sheet_info["sheet_name"]))
        return all_txs

    # =========================================================================
    # Private helpers
    # =========================================================================

    def _extract_rib(self, ws) -> str | None:
        """
        Extract the RIB from row 2 of a sheet.

        Row 2 contains: "R.I.B. : 10096 18027 00064764601"
        We return only the number part: "10096 18027 00064764601"
        """
        cell_value = ws["A2"].value
        if not cell_value:
            return None
        # Split on ": " — take the part after the colon
        if ": " in str(cell_value):
            return str(cell_value).split(": ", 1)[1].strip()
        return None

    def _extract_balance_from_sheet(self, ws) -> float | None:
        """
        Extract the closing balance from the footer row.

        Footer row (near the end of the sheet):
        col D = "Solde au 30/03/2026 : " or None
        col F = 798.27 (the balance value)
        """
        # Scan last 10 rows for the footer
        max_row = ws.max_row
        for row in ws.iter_rows(
            min_row=max(1, max_row - 10), max_row=max_row, values_only=True
        ):
            col_d = row[3] if len(row) > 3 else None
            col_f = row[5] if len(row) > 5 else None
            if isinstance(col_d, str) and "Solde au" in col_d:
                if isinstance(col_f, (int, float)):
                    return float(col_f)
        return None

    def _detect_account_type(self, ws) -> str:
        """
        Guess account type (checking or savings) from the title in row 1.

        Title examples:
          "Situation de votre compte C/C CONTRAT PERSONNEL GLOBAL ..."  → checking
          "Situation de votre compte LIVRET A SUP ..."                  → savings
          "Situation de votre compte LIVRET DE DEVELOPPEMENT DURABLE ..." → savings
        """
        title = str(ws["A1"].value or "").upper()
        if "C/C" in title or "CONTRAT PERSONNEL" in title:
            return "checking"
        return "savings"

    def _parse_row(
        self,
        date_val,  # datetime.datetime from openpyxl
        libelle: str,
        debit,  # float (negative) or None
        credit,  # float (positive) or None
        solde,  # float or None — running balance after this transaction
        currency: str,
        rib: str,  # account identifier — used in import_hash
        row_idx: int,  # kept for internal use, not in hash
        occurrence_counters: dict | None = None,  # mutated in place
    ) -> TransactionDict:
        """
        Convert one Excel row into a TransactionDict.

        Amount convention: negative = debit (money out), positive = credit (money in).
        CIC stores debit as negative float in column D, credit as positive in column E.

        occurrence_counters : {(rib, date, amount, desc) → next_index}.
        Permet de distinguer deux transactions identiques le même jour sans dépendre
        de la position dans le fichier (contrairement à row_idx qui changeait à chaque
        nouvel export et causait des doublons à la réimportation).
        """
        date_str = date_val.strftime("%Y-%m-%d")

        # Amount: one of the two columns is always None
        if debit is not None:
            amount = float(debit)  # already negative in the file
            curr = currency or "EUR"
        elif credit is not None:
            amount = float(credit)
            curr = currency or "EUR"
        else:
            raise ValueError(f"Both Débit and Crédit are None for row: {libelle}")

        description_raw = str(libelle).strip() if libelle else ""
        display_name = self._clean_merchant(description_raw)
        merchant_name = display_name  # pre-fill override
        card_last_four = self._parse_card(description_raw)

        # occurrence_index : 0 pour la 1ère occurrence de (rib, date, amount, desc)
        # dans ce fichier, 1 pour la 2ème, etc. Stable entre exports partiels et complets
        # car CIC préserve l'ordre intra-journalier. Inclus dans le hash pour éviter
        # les collisions sur des transactions genuinement distinctes mais identiques.
        if occurrence_counters is None:
            occurrence_counters = {}
        group_key = (rib, date_str, amount, description_raw)
        occurrence_index = occurrence_counters.get(group_key, 0)
        occurrence_counters[group_key] = occurrence_index + 1

        raw = f"{rib}|{date_str}|{amount}|{description_raw}|{occurrence_index}"
        import_hash = hashlib.sha256(raw.encode()).hexdigest()

        # Solde après transaction — colonne F de l'Excel CIC.
        # float si présent, None si la cellule est vide (rare mais possible).
        # Utilisé par ImportService pour créer des BalanceSnapshot journaliers.
        # Non utilisé dans le hash (voir contrat dans base.py).
        balance_after = float(solde) if isinstance(solde, (int, float)) else None

        return TransactionDict(
            date=date_str,
            time=None,  # CIC doesn't export transaction time
            amount=amount,
            currency=curr,
            description_raw=description_raw,
            display_name=display_name,
            merchant_name=merchant_name,
            card_last_four=card_last_four,
            import_hash=import_hash,
            balance_after=balance_after,
        )

    def _clean_merchant(self, description: str) -> str:
        """
        Remove structurally guaranteed CIC noise — nothing else.

        These two patterns are imposed by the French banking norm, never merchant names:
          - Verb prefix + optional DDMM date: "PAIEMENT PSC 1703 ", "VIR SEPA ", "RETRAIT DAB 2703 "
          - Card suffix: "CARTE 8703" and anything after (always terminal noise)

        Everything else (city names, reference codes, amounts) is left intact
        because it is ambiguous — removing it would lose information.
        """
        text = re.sub(
            r"^(?:PAIEMENT (?:PSC|CB) \d{4}|VIR (?:SEPA|INST|PERM)|RETRAIT DAB \d{4}) ",
            "",
            description,
        )
        text = re.sub(r"\s+CARTE \d{4}.*$", "", text)
        return self._normalize_merchant(text)

    def _parse_card(self, description: str) -> str | None:
        """
        Extract card last 4 digits from a CIC description.

        Format: "... CARTE 8703"  → "8703"
        Returns None if no CARTE reference found (transfers, fees...).
        """
        match = re.search(r"\bCARTE\s+(\d{4})\b", description, re.IGNORECASE)
        return match.group(1) if match else None
